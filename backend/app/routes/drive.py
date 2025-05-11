from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from tempfile import NamedTemporaryFile
from fpdf import FPDF
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fastapi.responses import FileResponse
from tempfile import NamedTemporaryFile
from typing import List
from sqlalchemy.orm import Session
from app.database import get_db
from app.dependencies import get_current_user
from app.schemas.drive_schema import VaccinationDriveCreate, VaccinationDriveUpdate, VaccinationDriveOut
from app.actions.vaccination_drive_actions import (
    create_vaccination_drive, list_vaccination_drives, get_vaccination_drive, update_vaccination_drive
)

router = APIRouter(prefix="/vaccination-drives", tags=["Vaccination Drives"])

# Route to create a new vaccination drive
@router.post("/")
def add_vaccination_drive(
    data: VaccinationDriveCreate, 
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    return create_vaccination_drive(db, data.vaccine_id, data.date, data.applicable_classes, data.available_doses)

# Route to get a list of all vaccination drives
@router.get("/", response_model=List[VaccinationDriveOut])
def get_all_vaccination_drives(
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    return list_vaccination_drives(db)

# Route to get a single vaccination drive by ID
@router.get("/{drive_id}")
def get_single_vaccination_drive(
    drive_id: int, 
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    drive = get_vaccination_drive(db, drive_id)
    if drive is None:
        raise HTTPException(status_code=404, detail="Vaccination drive not found")
    return drive

# Route to update an existing vaccination drive by ID
@router.put("/{drive_id}")
def update_vaccination_drive_details(
    drive_id: int, 
    data: VaccinationDriveUpdate,
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    return update_vaccination_drive(
        db, 
        drive_id, 
        data.date, 
        data.applicable_classes, 
        data.available_doses
    )

# Route to download vaccination drives as a CSV file
@router.get("/{drive_id}/export/csv")
def export_single_drive_csv(
    drive_id: int,
    db: Session = Depends(get_db)
):
    drive = get_vaccination_drive(db, drive_id)
    if not drive:
        raise HTTPException(status_code=404, detail="Vaccination drive not found")

    temp_file = NamedTemporaryFile(delete=False, suffix=".csv", mode="w", newline='', encoding='utf-8')
    writer = csv.writer(temp_file)
    writer.writerow([
        "Drive ID", "Vaccine", "Date", "Applicable Classes", "Available Doses", "Is Editable",
        "Student Name", "Student Class", "Student Unique ID", "Vaccination Date"
    ])

    if drive.vaccination_records:
        for record in drive.vaccination_records:
            writer.writerow([
                drive.id,
                record.vaccine.name,
                drive.date,
                drive.applicable_classes,
                drive.available_doses,
                drive.is_editable,
                record.student.name,
                record.student.student_class,
                record.student.unique_id,
                record.vaccination_date
            ])
    else:
        writer.writerow([
            drive.id,
            "",  # No vaccine details
            drive.date,
            drive.applicable_classes,
            drive.available_doses,
            drive.is_editable,
            "", "", "", ""
        ])

    temp_file.close()
    return FileResponse(path=temp_file.name, filename=f"vaccination_drive_{drive.id}.csv", media_type='text/csv')

# Route to download vaccination drives as a PDF file
@router.get("/{drive_id}/export/pdf")
def export_single_drive_pdf(
    drive_id: int,
    db: Session = Depends(get_db)
):
    drive = get_vaccination_drive(db, drive_id)
    if not drive:
        raise HTTPException(status_code=404, detail="Vaccination drive not found")

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt=f"Vaccination Drive ID: {drive.id}", ln=True)
    pdf.cell(200, 10, txt=f"Date: {drive.date}", ln=True)
    pdf.cell(200, 10, txt=f"Applicable Classes: {drive.applicable_classes}", ln=True)
    pdf.cell(200, 10, txt=f"Available Doses: {drive.available_doses}", ln=True)
    pdf.cell(200, 10, txt=f"Is Editable: {'Yes' if drive.is_editable else 'No'}", ln=True)
    pdf.cell(200, 10, txt="Vaccination Records:", ln=True)

    if drive.vaccination_records:
        for record in drive.vaccination_records:
            pdf.cell(200, 10, txt=f"- {record.student.name} ({record.student.student_class}) | "
                                  f"{record.vaccine.name} on {record.vaccination_date}", ln=True)
    else:
        pdf.cell(200, 10, txt="- No vaccination records", ln=True)

    temp_pdf = NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf.output(temp_pdf.name)
    return FileResponse(path=temp_pdf.name, filename=f"vaccination_drive_{drive.id}.pdf", media_type='application/pdf')

# Route to download vaccination drives as an Excel file
@router.get("/{drive_id}/export/excel")
def export_single_drive_excel(
    drive_id: int,
    db: Session = Depends(get_db)
):
    drive = get_vaccination_drive(db, drive_id)
    if not drive:
        raise HTTPException(status_code=404, detail="Vaccination drive not found")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Drive {drive.id}"

    # Write headers
    headers = [
        "Drive ID", "Vaccine", "Date", "Applicable Classes", "Available Doses", "Is Editable",
        "Student Name", "Student Class", "Student Unique ID", "Vaccination Date"
    ]
    ws.append(headers)

    # Write data rows
    if drive.vaccination_records:
        for record in drive.vaccination_records:
            ws.append([
                drive.id,
                record.vaccine.name,
                str(drive.date),
                drive.applicable_classes,
                drive.available_doses,
                "Yes" if drive.is_editable else "No",
                record.student.name,
                record.student.student_class,
                record.student.unique_id,
                str(record.vaccination_date)
            ])
    else:
        ws.append([
            drive.id, "", str(drive.date), drive.applicable_classes,
            drive.available_doses, "Yes" if drive.is_editable else "No",
            "", "", "", ""
        ])

    # Auto-width columns
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)

    return FileResponse(path=temp_file.name, filename=f"vaccination_drive_{drive.id}.xlsx", media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
